from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from Empresa import Empresa
# Inicializa o DataFrame vazio
df_empresas = pd.DataFrame(
    columns=["Link", "Reclamações Respondidas", "Voltariam ao Negócio", "Índice de Solução", "Nota do Consumidor"])


def adicionar_ao_dataframe(empresa_obj):
    global df_empresas

    empresa_dict = {
        "Link": empresa_obj.link,
        "Reclamações Respondidas": empresa_obj.reclamacoes_respondidas,
        "Voltariam ao Negócio": empresa_obj.voltariam_negocio,
        "Índice de Solução": empresa_obj.indice_solucao,
        "Nota do Consumidor": empresa_obj.nota_consumidor
    }
    novo_df = pd.DataFrame([empresa_dict])


    df_empresas = pd.concat([df_empresas, novo_df], ignore_index=True)


def pegar_melhores_empresas(driver):
    try:
        sleep(2)
        div = driver.find_element(By.XPATH, "/html/body/section[2]/div/astro-island/div/div[2]/div/div[1]")
        links = div.find_elements(By.TAG_NAME, 'a')
        for i, link in enumerate(links[:3]):
            print(f"Clicando no Link {i + 1}: {link.get_attribute('href')}")
            hrefs_melhores.append(link.get_attribute('href'))
            sleep(2)
        print(hrefs_melhores)

    except Exception as e:
        print(f"Erro ao encontrar ou clicar nos elementos: {e}")


def pegar_piores_empresas(driver):
    try:
        sleep(2)
        div = driver.find_element(By.XPATH, "/html/body/section[2]/div/astro-island/div/div[3]/div/div[2]")
        links = div.find_elements(By.TAG_NAME, 'a')
        for i, link in enumerate(links[:3]):
            print(f"Clicando no Link {i + 1}: {link.get_attribute('href')}")
            hrefs_piores.append(link.get_attribute('href'))
            sleep(2)
        print(hrefs_piores)

    except Exception as e:
        print(f"Erro ao encontrar ou clicar nos elementos: {e}")


def pintar_excel(wb):
    ws = wb.active

    # Definir as cores (verde claro e vermelho)
    verde_claro = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    vermelho = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

    # Aplica o verde claro nas 3 primeiras linhas (linhas 2, 3 e 4, já que a primeira linha é o cabeçalho)
    for row in range(2, 5):  # Linhas 2 a 4
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = verde_claro

    # Aplica o vermelho nas 3 últimas linhas
    for row in range(ws.max_row - 2, ws.max_row + 1):  # Últimas 3 linhas
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = vermelho

    # Salva o arquivo Excel com a formatação
    wb.save("empresas_dados_formatado.xlsx")
    print("Dados exportados com sucesso para 'empresas_dados.xlsx'")

# Inicializa o driver
options = Options()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

hrefs_melhores = []
hrefs_piores = []


driver.get('https://www.reclameaqui.com.br/')
try:


    element = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/a[1]"))

    )
    element.click()
    sleep(2)
    navegar = driver.find_element(By.XPATH, "/html/body/section[2]/div/astro-island/div/nav/button[2]")
    navegar.click()


    for i in range(5):
        navegar.click()
        print(f"Clique {i + 1} realizado")
        sleep(2)

    botao = driver.find_element(By.XPATH, "/html/body/section[2]/div/astro-island/div/nav/div[3]/button[9]")
    sleep(2)
    botao.click()

except Exception as e:
    print(f"Erro ao encontrar ou clicar no botão: {e}")


pegar_melhores_empresas(driver)
pegar_piores_empresas(driver)
driver.quit()


for url in hrefs_melhores:
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.get(url)
    sleep(5)
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')

    div_desejada = soup.find('div', class_='go267425901')
    textos_span = []
    divs_filho = div_desejada.findAll('div', class_='go4263471347')

    # Extraindo os dados e adicionando ao DataFrame
    for div in divs_filho:
        spans = div.findAll('span')
        textos_span.append(spans[0].text)

    reclamacoes_respondidas = textos_span[1]
    voltariam_negocio = textos_span[4]
    indice_solucao = textos_span[5]
    nota_consumidor = textos_span[3]
    textos_span.clear()
    # Cria o objeto da Empresa
    empresa = Empresa(link=url,
                      reclamacoes_respondidas=reclamacoes_respondidas,
                      voltariam_negocio=voltariam_negocio,
                      indice_solucao=indice_solucao,
                      nota_consumidor=nota_consumidor)

    # Adiciona ao DataFrame
    adicionar_ao_dataframe(empresa)
    driver.quit()

for url in hrefs_piores:
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.get(url)
    sleep(5)
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')

    div_desejada = soup.find('div', class_='go267425901')
    textos_span = []
    divs_filho = div_desejada.findAll('div', class_='go4263471347')

    # Extraindo os dados e adicionando ao DataFrame
    for div in divs_filho:
        spans = div.findAll('span')
        textos_span.append(spans[0].text)

    reclamacoes_respondidas = textos_span[1]
    voltariam_negocio = textos_span[4]
    indice_solucao = textos_span[5]
    nota_consumidor = textos_span[3]
    textos_span.clear()
    # Cria o objeto da Empresa
    empresa = Empresa(link=url,
                      reclamacoes_respondidas=reclamacoes_respondidas,
                      voltariam_negocio=voltariam_negocio,
                      indice_solucao=indice_solucao,
                      nota_consumidor=nota_consumidor)

    # Adiciona ao DataFrame
    adicionar_ao_dataframe(empresa)
    driver.quit()

df_empresas.to_excel("empresas_dados.xlsx", index=False)

wb = load_workbook("empresas_dados.xlsx")
pintar_excel(wb)